"""STA Excel importer — enriches existing RTM projects with Software Traceability Analysis data.

Parses two sheets from the STA Excel:
  - "SW Trace Analysis": SRS spec refs, per-module evidence, design outputs, unit tests
  - "Design Input Verification": version-by-version test evidence history

Column mapping for "SW Trace Analysis":
     0: Req IDs (SRD-xxx)               8: PCU evidence
     1: PRD Section Header              9: LVP evidence
     2: Description                    10: SYR evidence
     3: Modules                        11: PCA evidence
     4: ASWS spec ref                  12: EtCO2 evidence
     5: ASWUI spec ref                 13: AutoID evidence
     6: ASWIS spec ref                 14: Design Outputs (SDD refs)
     7: Total Test Evidence            15: Unit Test Verification

Column mapping for "Design Input Verification":
     0: Source ID (SRD-xxx)
     1-2: v12.6 Test Ref + TER
     3-4: v12.4.0 Test Ref + TER
     ... (7 version pairs total)
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Optional

from core.shared.database import SharedDatabase


# ── SW Trace Analysis column indices ──────────────────────────────

STA_COL_SRD_ID = 0
STA_COL_PRD_SECTION = 1
STA_COL_DESCRIPTION = 2
STA_COL_MODULES = 3
STA_COL_ASWS = 4
STA_COL_ASWUI = 5
STA_COL_ASWIS = 6
STA_COL_TOTAL_EVIDENCE = 7
STA_COL_PCU = 8
STA_COL_LVP = 9
STA_COL_SYR = 10
STA_COL_PCA = 11
STA_COL_ETCO2 = 12
STA_COL_AUTOID = 13
STA_COL_DESIGN_OUTPUTS = 14
STA_COL_UNIT_TESTS = 15

STA_MODULE_EVIDENCE_COLS = [
    (STA_COL_PCU, "PCU"),
    (STA_COL_LVP, "LVP"),
    (STA_COL_SYR, "SYR"),
    (STA_COL_PCA, "PCA"),
    (STA_COL_ETCO2, "EtCO2"),
    (STA_COL_AUTOID, "Auto-ID"),
]

# ── Design Input Verification column mapping ──────────────────────

DIV_COL_SRD_ID = 0
DIV_VERSION_COLS = [
    (1, 2, "v12.6"),
    (3, 4, "v12.4.0"),
    (5, 6, "v12.3.2"),
    (7, 8, "v12.3.1"),
    (9, 10, "v12.3"),
    (11, 12, "v12.2"),
    (13, 14, "v12.1.x"),
]


# ── Helpers ───────────────────────────────────────────────────────

def _cell_str(row: tuple, idx: int) -> str:
    if idx >= len(row) or row[idx] is None:
        return ""
    val = str(row[idx]).strip()
    if val.upper() == "N/A":
        return ""
    return val


def _parse_comma_list(raw: str) -> list[str]:
    """Parse comma-separated values, filtering blanks and N/A."""
    if not raw:
        return []
    items = []
    for part in raw.split(","):
        part = part.strip()
        if part and part.upper() != "N/A":
            items.append(part)
    return items


def _parse_tc_ids(raw: str) -> list[str]:
    """Parse TC IDs from comma/newline-separated cells."""
    if not raw:
        return []
    ids = []
    for line in raw.replace("\n", ",").split(","):
        line = line.strip()
        if not line or line.upper() == "N/A":
            continue
        ids.append(line)
    return ids


def _detect_sheet(sheetnames: list[str], keywords: list[str]) -> Optional[str]:
    """Find a sheet whose name contains any of the keywords (case-insensitive)."""
    for name in sheetnames:
        lower = name.lower()
        if any(kw in lower for kw in keywords):
            return name
    return None


# ── Main import function ──────────────────────────────────────────

def import_sta_excel(
    file_path: str,
    project_id: int,
    sw_trace_sheet: str | None = None,
    div_sheet: str | None = None,
    db: SharedDatabase | None = None,
) -> dict:
    """Import STA Excel and enrich an existing RTM project.

    Args:
        file_path: Path to STA .xlsx file
        project_id: Existing RTM project ID to enrich
        sw_trace_sheet: Sheet name for SW Trace Analysis (auto-detects)
        div_sheet: Sheet name for Design Input Verification (auto-detects)
        db: SharedDatabase instance (uses singleton if None)

    Returns:
        Import stats with matched/unmatched counts per sheet.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    db = db or SharedDatabase.get_instance()
    file_path = str(Path(file_path).resolve())

    # Validate project exists
    project = db.fetchone("SELECT id FROM rtm_projects WHERE id = ?", (project_id,))
    if not project:
        raise ValueError(f"Project {project_id} not found")

    # Build SRD ID -> requirement IDs lookup (one SRD can map to multiple reqs)
    rows = db.fetchall(
        "SELECT id, srd_id FROM rtm_requirements WHERE project_id = ?",
        (project_id,),
    )
    srd_lookup: dict[str, list[int]] = {}
    for r in rows:
        srd = r["srd_id"].strip()
        if srd:
            srd_lookup.setdefault(srd, []).append(r["id"])

    # Clear previous STA data for this project (idempotent re-import)
    for table in [
        "rtm_sta_spec_refs",
        "rtm_sta_module_evidence",
        "rtm_sta_design_outputs",
        "rtm_sta_version_verification",
    ]:
        db.execute(f"DELETE FROM {table} WHERE project_id = ?", (project_id,))
    db.commit()

    # Load workbook
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    # Auto-detect sheets
    if sw_trace_sheet is None:
        sw_trace_sheet = _detect_sheet(wb.sheetnames, ["trace analysis", "sw trace"])
    if div_sheet is None:
        div_sheet = _detect_sheet(
            wb.sheetnames, ["design input", "verification"]
        )

    result = {
        "status": "success",
        "project_id": project_id,
        "sw_trace": None,
        "design_input_verification": None,
    }

    # ── Process SW Trace Analysis ─────────────────────────────────
    if sw_trace_sheet and sw_trace_sheet in wb.sheetnames:
        result["sw_trace"] = _import_sw_trace(
            wb[sw_trace_sheet], project_id, srd_lookup, db
        )

    # ── Process Design Input Verification ─────────────────────────
    if div_sheet and div_sheet in wb.sheetnames:
        result["design_input_verification"] = _import_div(
            wb[div_sheet], project_id, srd_lookup, db
        )

    # Update project enrichment timestamp
    db.execute(
        "UPDATE rtm_projects SET sta_enriched_at = datetime('now') WHERE id = ?",
        (project_id,),
    )
    db.commit()
    wb.close()

    return result


def _import_sw_trace(
    ws, project_id: int, srd_lookup: dict[str, list[int]], db: SharedDatabase
) -> dict:
    """Process the SW Trace Analysis sheet."""
    stats = {
        "rows_processed": 0,
        "matched": 0,
        "unmatched": 0,
        "unmatched_srd_ids": [],
        "spec_refs_created": 0,
        "module_evidence_created": 0,
        "design_outputs_created": 0,
    }

    header_rows = 2
    row_num = 0

    for row in ws.iter_rows(max_col=16, values_only=True):
        row_num += 1
        if row_num <= header_rows:
            continue

        srd_id = _cell_str(row, STA_COL_SRD_ID)
        if not srd_id:
            continue

        stats["rows_processed"] += 1
        req_ids = srd_lookup.get(srd_id)

        if not req_ids:
            stats["unmatched"] += 1
            if len(stats["unmatched_srd_ids"]) < 50:
                stats["unmatched_srd_ids"].append(srd_id)
            continue

        stats["matched"] += 1
        prd_section = _cell_str(row, STA_COL_PRD_SECTION)
        description = _cell_str(row, STA_COL_DESCRIPTION)
        modules = _cell_str(row, STA_COL_MODULES)
        asws = _cell_str(row, STA_COL_ASWS)
        aswui = _cell_str(row, STA_COL_ASWUI)
        aswis = _cell_str(row, STA_COL_ASWIS)

        for req_id in req_ids:
            if asws or aswui or aswis:
                db.execute(
                    "INSERT INTO rtm_sta_spec_refs "
                    "(requirement_id, project_id, srd_id, prd_section, description, "
                    " sta_modules, asws_ref, aswui_ref, aswis_ref) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (req_id, project_id, srd_id, prd_section, description,
                     modules, asws, aswui, aswis),
                )
                stats["spec_refs_created"] += 1

            total_ev = _cell_str(row, STA_COL_TOTAL_EVIDENCE)
            if total_ev:
                tc_ids = _parse_tc_ids(total_ev)
                if tc_ids:
                    db.execute(
                        "INSERT INTO rtm_sta_module_evidence "
                        "(requirement_id, project_id, srd_id, module_name, "
                        " tc_ids_json, tc_count) "
                        "VALUES (?, ?, ?, ?, ?, ?)",
                        (req_id, project_id, srd_id, "TOTAL",
                         json.dumps(tc_ids), len(tc_ids)),
                    )
                    stats["module_evidence_created"] += 1

            for col_idx, mod_name in STA_MODULE_EVIDENCE_COLS:
                raw = _cell_str(row, col_idx)
                if not raw:
                    continue
                tc_ids = _parse_tc_ids(raw)
                if tc_ids:
                    db.execute(
                        "INSERT INTO rtm_sta_module_evidence "
                        "(requirement_id, project_id, srd_id, module_name, "
                        " tc_ids_json, tc_count) "
                        "VALUES (?, ?, ?, ?, ?, ?)",
                        (req_id, project_id, srd_id, mod_name,
                         json.dumps(tc_ids), len(tc_ids)),
                    )
                    stats["module_evidence_created"] += 1

            design_raw = _cell_str(row, STA_COL_DESIGN_OUTPUTS)
            ut_raw = _cell_str(row, STA_COL_UNIT_TESTS)
            design_refs = _parse_comma_list(design_raw)
            ut_files = _parse_comma_list(ut_raw)

            if design_refs or ut_files:
                db.execute(
                    "INSERT INTO rtm_sta_design_outputs "
                    "(requirement_id, project_id, srd_id, design_refs_json, "
                    " design_ref_count, unit_test_files_json, unit_test_count) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (req_id, project_id, srd_id,
                     json.dumps(design_refs), len(design_refs),
                     json.dumps(ut_files), len(ut_files)),
                )
                stats["design_outputs_created"] += 1

        if stats["rows_processed"] % 200 == 0:
            db.commit()

    db.commit()
    return stats


def _import_div(
    ws, project_id: int, srd_lookup: dict[str, list[int]], db: SharedDatabase
) -> dict:
    """Process the Design Input Verification sheet."""
    stats = {
        "rows_processed": 0,
        "matched": 0,
        "unmatched": 0,
        "version_records_created": 0,
    }

    header_rows = 2
    row_num = 0

    for row in ws.iter_rows(max_col=15, values_only=True):
        row_num += 1
        if row_num <= header_rows:
            continue

        srd_id = _cell_str(row, DIV_COL_SRD_ID)
        if not srd_id:
            continue

        stats["rows_processed"] += 1
        req_ids = srd_lookup.get(srd_id)

        if not req_ids:
            stats["unmatched"] += 1
            continue

        stats["matched"] += 1

        for req_id in req_ids:
            for test_col, ter_col, version_label in DIV_VERSION_COLS:
                test_ref = _cell_str(row, test_col)
                ter_ref = _cell_str(row, ter_col)

                if not test_ref and not ter_ref:
                    continue

                db.execute(
                    "INSERT INTO rtm_sta_version_verification "
                    "(requirement_id, project_id, srd_id, version_label, "
                    " test_ref, ter_ref) "
                    "VALUES (?, ?, ?, ?, ?, ?)",
                    (req_id, project_id, srd_id, version_label,
                     test_ref, ter_ref),
                )
                stats["version_records_created"] += 1

        if stats["rows_processed"] % 500 == 0:
            db.commit()

    db.commit()
    return stats
