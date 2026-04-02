"""Resource Excel importer — parses FY26 Infusion resources.xlsm into the shared DB.

Uses the $PivotView sheet (normalized: one row per person-per-month).
Columns: HC, Type, Team, Location, Line Manager, Name, Project, Activity,
         Comment, Quartile (datetime), Allocation (float 0-1).
"""
from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path

from core.shared.database import SharedDatabase


def import_resource_excel(
    file_path: str,
    sheet_name: str | None = None,
    db: SharedDatabase | None = None,
) -> dict:
    """Import a resource planning Excel file.

    Returns dict with import stats.
    """
    import openpyxl

    db = db or SharedDatabase.get_instance()
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        # Prefer $PivotView (normalized), fall back to Resources sheet
        target = sheet_name
        if not target:
            for name in ("$PivotView", "PivotView", "Resources"):
                if name in wb.sheetnames:
                    target = name
                    break
        if not target:
            target = wb.sheetnames[0]

        ws = wb[target]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if len(rows) < 2:
        raise ValueError("Sheet has no data rows")

    headers = [str(h).strip() if h else "" for h in rows[0]]

    # Map headers to indices (case-insensitive)
    hmap = {h.lower(): i for i, h in enumerate(headers)}

    def col(name, *alts):
        for n in (name, *alts):
            if n.lower() in hmap:
                return hmap[n.lower()]
        return None

    # Detect if this is the pivoted sheet (has Quartile+Allocation)
    # or the wide sheet (has monthly date columns)
    is_pivot = col("quartile") is not None and col("allocation") is not None

    # Clear existing resource data
    db.execute("DELETE FROM resource_allocations")
    db.execute("DELETE FROM resource_people")
    db.commit()

    people_seen = {}
    alloc_rows = []

    if is_pivot:
        # Normalized: one row per person-per-month
        ci = {
            "name": col("name"),
            "type": col("type", "hc"),
            "team": col("team"),
            "location": col("location"),
            "manager": col("line manager", "manager"),
            "project": col("project"),
            "activity": col("activity"),
            "comment": col("comment"),
            "quartile": col("quartile"),
            "allocation": col("allocation"),
        }

        for row in rows[1:]:
            name = str(row[ci["name"]] or "").strip()
            if not name:
                continue

            # Upsert person
            if name not in people_seen:
                people_seen[name] = {
                    "name": name,
                    "type": str(row[ci["type"]] or "FTE").strip(),
                    "team": str(row[ci["team"]] or "").strip(),
                    "location": str(row[ci["location"]] or "").strip(),
                    "manager": str(row[ci["manager"]] or "").strip() if ci["manager"] is not None else "",
                    "project": str(row[ci["project"]] or "").strip(),
                    "activity": str(row[ci["activity"]] or "").strip(),
                }

            # Parse month
            q = row[ci["quartile"]]
            if isinstance(q, datetime):
                month = q.strftime("%Y-%m")
            elif q:
                month = str(q)[:7]
            else:
                continue

            alloc_val = float(row[ci["allocation"]] or 0)
            alloc_rows.append((name, row[ci["project"]] or "", month, alloc_val))

    else:
        # Wide format: monthly columns as dates
        ci_name = col("name")
        ci_type = col("type", "hc")
        ci_team = col("team")
        ci_loc = col("location")
        ci_mgr = col("line manager", "manager")
        ci_proj = col("project")
        ci_act = col("activity")
        ci_comment = col("comment")

        month_cols = []
        for i, h in enumerate(headers):
            if h and len(h) >= 6:
                try:
                    if isinstance(rows[0][i], datetime):
                        month_cols.append((i, rows[0][i].strftime("%Y-%m")))
                    elif "-" in h and len(h) == 7:
                        month_cols.append((i, h))
                except (ValueError, TypeError):
                    pass

        start_row = 1
        if ci_name is None:
            headers2 = [str(h).strip() if h else "" for h in rows[2]] if len(rows) > 2 else []
            hmap2 = {h.lower(): i for i, h in enumerate(headers2)}
            ci_name = hmap2.get("name")
            ci_type = hmap2.get("type", hmap2.get("hc"))
            ci_team = hmap2.get("team")
            ci_loc = hmap2.get("location")
            ci_mgr = hmap2.get("line manager", hmap2.get("manager"))
            ci_proj = hmap2.get("project")
            ci_act = hmap2.get("activity")
            start_row = 3

        if ci_name is None:
            return {"status": "error", "detail": "Could not find 'Name' column"}

        for row in rows[start_row:]:
            name = str(row[ci_name] or "").strip() if ci_name is not None else ""
            if not name:
                continue

            if name not in people_seen:
                people_seen[name] = {
                    "name": name,
                    "type": str(row[ci_type] or "FTE").strip() if ci_type is not None else "FTE",
                    "team": str(row[ci_team] or "").strip() if ci_team is not None else "",
                    "location": str(row[ci_loc] or "").strip() if ci_loc is not None else "",
                    "manager": str(row[ci_mgr] or "").strip() if ci_mgr is not None else "",
                    "project": str(row[ci_proj] or "").strip() if ci_proj is not None else "",
                    "activity": str(row[ci_act] or "").strip() if ci_act is not None else "",
                }

            project = str(row[ci_proj] or "").strip() if ci_proj is not None else ""
            for mc_idx, mc_month in month_cols:
                val = row[mc_idx]
                alloc_val = float(val) if val else 0.0
                alloc_rows.append((name, project, mc_month, alloc_val))

    try:
        # Insert people
        for p in people_seen.values():
            db.execute(
                "INSERT INTO resource_people (name, type, team, location, manager, project, activity) "
                "VALUES (?, ?, ?, ?, ?, ?, ?)",
                (p["name"], p["type"], p["team"], p["location"], p["manager"], p["project"], p["activity"]),
            )

        # Insert allocations
        if alloc_rows:
            db.executemany(
                "INSERT INTO resource_allocations (person_name, project, month, allocation) "
                "VALUES (?, ?, ?, ?)",
                alloc_rows,
            )

        db.commit()
    except Exception:
        db.rollback()
        raise

    return {
        "status": "success",
        "people_imported": len(people_seen),
        "allocation_rows": len(alloc_rows),
        "sheet_used": target,
    }
