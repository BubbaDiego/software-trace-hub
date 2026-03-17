"""RTM Excel importer — parses Alaris-style RTM spreadsheets into rtm.db.

Column mapping (from 'RTM - Alaris v12.6 510k' and similar sheets):
    0: SNO                        15: SPEC_TCsManual
    1: xx (composite key)         16: REQ_TCsManual
    2: xxx                        17: MODULE_LVP_Manual
    3: SoftwareFeature            18: MODULE_SYR_Manual
    4: Feature                    19: MODULE_PCA_Manual
    5: Sub-Feature                20: MODULE_PCU
    6: SoftwareFunction           21: MODULE_EtCO2
    7: TrackerID                  22: Auto-ID
    8: HazardId                   23: MODULE_SPO2
    9: PCUREQSRDID                24: ScenarioBasedtestevidence(Manual)
   10: ImpactedModule             25: SPEC_TCsCATS
   11: PRDModules                 26: REQ_TCsCATS
   12: PCUREQSRSID                27: MODULE_LVPCATS
   13: PCUSPECID                  28: MODULE_SYRCATS
   14: Project12.3 (X;R;N)        29: MODULE_PCACATS
                                  30: ScenarioBasedtestevidence(CATS)
                                  31: Comment
"""
from __future__ import annotations

import hashlib
import json
import os
from pathlib import Path
from typing import Optional

from .rtm_db import RTMDatabase
from .models import EvidenceType, TraceStatus


# Column indices
COL_SNO = 0
COL_COMPOSITE = 1
COL_SW_FEATURE = 3
COL_FEATURE = 4
COL_SUB_FEATURE = 5
COL_SW_FUNCTION = 6
COL_TRACKER_ID = 7
COL_HAZARD_ID = 8
COL_SRD_ID = 9
COL_IMPACTED_MODULES = 10
COL_PRD_MODULES = 11
COL_SRS_ID = 12
COL_SPEC_ID = 13
COL_PROJECT_STATUS = 14
COL_SPEC_TCS_MANUAL = 15
COL_REQ_TCS_MANUAL = 16
COL_MODULE_LVP_MANUAL = 17
COL_MODULE_SYR_MANUAL = 18
COL_MODULE_PCA_MANUAL = 19
COL_MODULE_PCU_MANUAL = 20
COL_MODULE_ETCO2_MANUAL = 21
COL_MODULE_AUTOID_MANUAL = 22
COL_MODULE_SPO2_MANUAL = 23
COL_SCENARIO_MANUAL = 24
COL_SPEC_TCS_CATS = 25
COL_REQ_TCS_CATS = 26
COL_MODULE_LVP_CATS = 27
COL_MODULE_SYR_CATS = 28
COL_MODULE_PCA_CATS = 29
COL_SCENARIO_CATS = 30
COL_COMMENT = 31

# Evidence column → (EvidenceType, module_name)
EVIDENCE_MAP = [
    (COL_SPEC_TCS_MANUAL,     EvidenceType.SPEC_MANUAL,          ""),
    (COL_REQ_TCS_MANUAL,      EvidenceType.REQ_MANUAL,           ""),
    (COL_MODULE_LVP_MANUAL,   EvidenceType.MODULE_LVP_MANUAL,    "LVP"),
    (COL_MODULE_SYR_MANUAL,   EvidenceType.MODULE_SYR_MANUAL,    "SYR"),
    (COL_MODULE_PCA_MANUAL,   EvidenceType.MODULE_PCA_MANUAL,    "PCA"),
    (COL_MODULE_PCU_MANUAL,   EvidenceType.MODULE_PCU_MANUAL,    "PCU"),
    (COL_MODULE_ETCO2_MANUAL, EvidenceType.MODULE_ETCO2_MANUAL,  "EtCO2"),
    (COL_MODULE_AUTOID_MANUAL,EvidenceType.MODULE_AUTOID_MANUAL, "Auto-ID"),
    (COL_MODULE_SPO2_MANUAL,  EvidenceType.MODULE_SPO2_MANUAL,   "SPO2"),
    (COL_SPEC_TCS_CATS,       EvidenceType.CATS_SPEC,            ""),
    (COL_REQ_TCS_CATS,        EvidenceType.CATS_REQ,             ""),
    (COL_MODULE_LVP_CATS,     EvidenceType.CATS_MODULE_LVP,      "LVP"),
    (COL_MODULE_SYR_CATS,     EvidenceType.CATS_MODULE_SYR,      "SYR"),
    (COL_MODULE_PCA_CATS,     EvidenceType.CATS_MODULE_PCA,      "PCA"),
    (COL_SCENARIO_MANUAL,     EvidenceType.SCENARIO_MANUAL,      ""),
    (COL_SCENARIO_CATS,       EvidenceType.SCENARIO_CATS,        ""),
]


def _cell_str(row: tuple, idx: int) -> str:
    """Safely extract a cell value as a stripped string."""
    if idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).strip()


def _cell_int(row: tuple, idx: int) -> Optional[int]:
    """Safely extract a cell as int, or None."""
    val = _cell_str(row, idx)
    if not val:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _parse_tc_ids(raw: str) -> list[str]:
    """Parse a cell containing comma-separated TC IDs or 'SRD-XXXX [TXXXXXX]' patterns.

    Handles formats:
        '55771, 60430, 60431'
        'SRD-2420 [T000909], SRD-2420 [T000905]'
        Mixed/messy whitespace
    """
    if not raw:
        return []

    ids = []
    for part in raw.split(","):
        part = part.strip()
        if not part:
            continue
        # Handle 'SRD-XXXX [TXXXXXX]' format — extract the T-code
        if "[" in part and "]" in part:
            inner = part[part.index("[") + 1:part.index("]")].strip()
            if inner:
                ids.append(inner)
        else:
            ids.append(part)
    return ids


def _compute_trace_status(has_manual: bool, has_cats: bool, has_spec_id: bool) -> str:
    """Determine trace status for a requirement."""
    if has_manual and has_cats:
        return TraceStatus.FULL.value
    if has_manual or has_cats:
        return TraceStatus.PARTIAL.value
    return TraceStatus.MISSING.value


def compute_file_hash(file_path: str) -> str:
    """SHA-256 hash of the file for dedup."""
    h = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def import_excel(
    file_path: str,
    sheet_name: str | None = None,
    project_name: str = "",
    project_version: str = "",
    db: RTMDatabase | None = None,
) -> dict:
    """Import an RTM Excel file into the database.

    Args:
        file_path: Path to .xlsx file
        sheet_name: Sheet to import (auto-detects '510k' sheet if None)
        project_name: Display name for this import
        project_version: Version string (e.g. 'v12.6')
        db: RTMDatabase instance (uses singleton if None)

    Returns:
        dict with import stats: project_id, requirements, evidence_rows, etc.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    db = db or RTMDatabase.get_instance()
    file_path = str(Path(file_path).resolve())

    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check for duplicate import
    file_hash = compute_file_hash(file_path)
    existing = db.fetchone(
        "SELECT id FROM rtm_projects WHERE file_hash = ?", (file_hash,)
    )
    if existing:
        return {
            "status": "duplicate",
            "project_id": existing["id"],
            "message": f"File already imported as project {existing['id']}",
        }

    # Load workbook
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    # Auto-detect sheet
    if sheet_name is None:
        for name in wb.sheetnames:
            if "510k" in name.lower() or "v12.6" in name.lower():
                sheet_name = name
                break
        if sheet_name is None:
            sheet_name = wb.sheetnames[0]

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        )

    ws = wb[sheet_name]

    # Derive project name/version from sheet if not provided
    if not project_name:
        project_name = sheet_name
    if not project_version:
        for token in sheet_name.replace("-", " ").split():
            if token.startswith("v") and any(c.isdigit() for c in token):
                project_version = token
                break

    # Create project record
    db.execute(
        "INSERT INTO rtm_projects (name, version, sheet_name, file_hash) "
        "VALUES (?, ?, ?, ?)",
        (project_name, project_version, sheet_name, file_hash),
    )
    db.commit()
    project_id = db.fetchone(
        "SELECT id FROM rtm_projects WHERE file_hash = ?", (file_hash,)
    )["id"]

    # Parse rows
    req_count = 0
    evidence_count = 0
    skipped = 0
    first_row = True

    for row in ws.iter_rows(max_col=32, values_only=True):
        # Skip header row
        if first_row:
            first_row = False
            continue

        # Skip empty rows
        sno = _cell_int(row, COL_SNO)
        srd_id = _cell_str(row, COL_SRD_ID)
        feature = _cell_str(row, COL_FEATURE)

        if sno is None and not srd_id and not feature:
            skipped += 1
            continue

        # Insert requirement
        db.execute(
            "INSERT INTO rtm_requirements "
            "(project_id, sno, composite_key, software_feature, feature, "
            " sub_feature, software_function, tracker_id, hazard_id, "
            " srd_id, impacted_modules, prd_modules, srs_id, spec_id, "
            " project_status, comment, trace_status) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                project_id,
                sno,
                _cell_str(row, COL_COMPOSITE),
                _cell_str(row, COL_SW_FEATURE),
                feature,
                _cell_str(row, COL_SUB_FEATURE),
                _cell_str(row, COL_SW_FUNCTION),
                _cell_str(row, COL_TRACKER_ID),
                _cell_str(row, COL_HAZARD_ID),
                srd_id,
                _cell_str(row, COL_IMPACTED_MODULES),
                _cell_str(row, COL_PRD_MODULES),
                _cell_str(row, COL_SRS_ID),
                _cell_str(row, COL_SPEC_ID),
                _cell_str(row, COL_PROJECT_STATUS),
                _cell_str(row, COL_COMMENT),
                TraceStatus.MISSING.value,  # updated below
            ),
        )
        req_id = db.fetchone("SELECT last_insert_rowid() as id")["id"]
        req_count += 1

        # Parse evidence columns
        has_manual = False
        has_cats = False

        for col_idx, ev_type, module_name in EVIDENCE_MAP:
            raw = _cell_str(row, col_idx)
            tc_ids = _parse_tc_ids(raw)
            if not tc_ids:
                continue

            db.execute(
                "INSERT INTO rtm_test_evidence "
                "(requirement_id, evidence_type, module_name, tc_ids_json, tc_count) "
                "VALUES (?, ?, ?, ?, ?)",
                (req_id, ev_type.value, module_name, json.dumps(tc_ids), len(tc_ids)),
            )
            evidence_count += 1

            # Track manual vs CATS
            if "cats" in ev_type.value or "scenario_cats" == ev_type.value:
                has_cats = True
            else:
                has_manual = True

        # Update trace status
        has_spec = bool(_cell_str(row, COL_SPEC_ID))
        trace = _compute_trace_status(has_manual, has_cats, has_spec)
        db.execute(
            "UPDATE rtm_requirements SET trace_status = ? WHERE id = ?",
            (trace, req_id),
        )

        # Batch commit every 500 rows
        if req_count % 500 == 0:
            db.commit()

    # Final commit
    db.execute(
        "UPDATE rtm_projects SET total_requirements = ? WHERE id = ?",
        (req_count, project_id),
    )
    db.commit()
    wb.close()

    return {
        "status": "success",
        "project_id": project_id,
        "sheet_name": sheet_name,
        "requirements_imported": req_count,
        "evidence_rows_created": evidence_count,
        "rows_skipped": skipped,
    }


def import_all_sheets(
    file_path: str,
    db: RTMDatabase | None = None,
) -> list[dict]:
    """Import all RTM-like sheets from an Excel file.

    Skips sheets that don't look like RTM data (no SNO/SRD columns).
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    db = db or RTMDatabase.get_instance()
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Check if first row looks like RTM headers
        first_row = next(ws.iter_rows(max_col=15, max_row=1, values_only=True), None)
        if first_row is None:
            continue

        headers = [str(c).lower() if c else "" for c in first_row]
        has_sno = any("sno" in h for h in headers)
        has_srd = any("srd" in h or "reqsrd" in h for h in headers)

        if not (has_sno or has_srd):
            results.append({
                "sheet_name": sheet_name,
                "status": "skipped",
                "reason": "no RTM headers detected",
            })
            continue

        try:
            result = import_excel(
                file_path,
                sheet_name=sheet_name,
                db=db,
            )
            results.append(result)
        except Exception as e:
            results.append({
                "sheet_name": sheet_name,
                "status": "error",
                "message": str(e),
            })

    wb.close()
    return results
