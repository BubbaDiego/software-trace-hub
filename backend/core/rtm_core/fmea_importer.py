"""FMEA Excel importer — parses Alaris-style SW FMEA spreadsheets into rtm.db.

Uses the dFMEA sheet (main data). Has a 3-row merged header:
  Row 1: group headers (Risk Identification, Risk Controls, etc.)
  Row 2-3: column-level headers

Key columns mapped by position:
  A: FMEA ID, B: Source, C: Hazard, D: Hazardous Situation,
  E: Product(s), F: System/Sub-Assembly, G: Design Element/Component,
  H: Critical Component, I: Function, J: Failure Mode,
  K: Failure Code, L: Failure Code Description, M: Cause(s),
  N: Effect, O: System Effect,
  P: Risk Control Measures, Q: RCM Type, R: Corresponding Requirement,
  S: UI Specification, T: Evidence/Traceability, U: Standard Reference,
  V-Z: Probability estimates, AA: Mitigation Rationale,
  AB-AH: Residual risk columns
"""
from __future__ import annotations

from .rtm_db import RTMDatabase


def import_fmea_excel(
    file_path: str,
    sheet_name: str | None = None,
    db: RTMDatabase | None = None,
) -> dict:
    """Import an FMEA Excel file.

    Returns dict with import stats.
    """
    import openpyxl

    db = db or RTMDatabase.get_instance()
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    # Find the dFMEA sheet
    target = sheet_name
    if not target:
        for name in ("dFMEA", "FMEA", "SW FMEA"):
            if name in wb.sheetnames:
                target = name
                break
    if not target:
        # Try to find any sheet with "FMEA" in the name
        for name in wb.sheetnames:
            if "fmea" in name.lower():
                target = name
                break
    if not target:
        wb.close()
        return {"status": "error", "detail": f"No FMEA sheet found. Sheets: {wb.sheetnames}"}

    ws = wb[target]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 4:
        return {"status": "error", "detail": "Sheet has fewer than 4 rows (need header + data)"}

    # Data starts at row 4 (index 3) — first 3 rows are merged headers
    data_start = 3

    # Try to detect header row by looking for "FMEA" in first column
    for i, row in enumerate(rows[:5]):
        val = str(row[0] or "").strip().lower()
        if val in ("fmea id", "fmea_id", "id", "fmea"):
            data_start = i + 1
            break

    # Clear existing FMEA data
    db.execute("DELETE FROM fmea_common_causes")
    db.execute("DELETE FROM fmea_records")
    db.commit()

    def s(val):
        """Safely convert cell to string."""
        if val is None:
            return ""
        return str(val).strip()

    records = []
    for row in rows[data_start:]:
        fmea_id = s(row[0]) if len(row) > 0 else ""
        if not fmea_id:
            continue

        record = (
            fmea_id,
            s(row[1]) if len(row) > 1 else "",   # source
            s(row[2]) if len(row) > 2 else "",   # hazard
            s(row[3]) if len(row) > 3 else "",   # hazardous_situation
            s(row[4]) if len(row) > 4 else "",   # product
            s(row[5]) if len(row) > 5 else "",   # system
            s(row[6]) if len(row) > 6 else "",   # component
            s(row[7]) if len(row) > 7 else "",   # critical_component
            s(row[8]) if len(row) > 8 else "",   # function
            s(row[9]) if len(row) > 9 else "",   # failure_mode
            s(row[10]) if len(row) > 10 else "",  # failure_code
            s(row[11]) if len(row) > 11 else "",  # failure_code_desc
            s(row[12]) if len(row) > 12 else "",  # cause
            s(row[13]) if len(row) > 13 else "",  # effect
            s(row[14]) if len(row) > 14 else "",  # system_effect
            s(row[15]) if len(row) > 15 else "",  # rcm
            s(row[16]) if len(row) > 16 else "",  # rcm_type
            s(row[17]) if len(row) > 17 else "",  # requirement
            s(row[18]) if len(row) > 18 else "",  # ui_spec
            s(row[19]) if len(row) > 19 else "",  # evidence
            s(row[20]) if len(row) > 20 else "",  # standard_ref
            s(row[21]) if len(row) > 21 else "",  # p1a
            s(row[22]) if len(row) > 22 else "",  # p1b_p1c
            s(row[23]) if len(row) > 23 else "",  # p1_sources
            s(row[24]) if len(row) > 24 else "",  # p1
            s(row[25]) if len(row) > 25 else "",  # poh
            s(row[26]) if len(row) > 26 else "",  # mitigation_rationale
            s(row[27]) if len(row) > 27 else "",  # residual_p1a
            s(row[28]) if len(row) > 28 else "",  # residual_p1b_p1c
            s(row[29]) if len(row) > 29 else "",  # residual_p1_sources
            s(row[30]) if len(row) > 30 else "",  # residual_p1
            s(row[31]) if len(row) > 31 else "",  # residual_hazardous_situation
            s(row[32]) if len(row) > 32 else "",  # severity
            s(row[33]) if len(row) > 33 else "",  # residual_poh
        )
        records.append(record)

    if records:
        db.executemany(
            """INSERT INTO fmea_records (
                fmea_id, source, hazard, hazardous_situation, product,
                system, component, critical_component, function, failure_mode,
                failure_code, failure_code_desc, cause, effect, system_effect,
                rcm, rcm_type, requirement, ui_spec, evidence, standard_ref,
                p1a, p1b_p1c, p1_sources, p1, poh, mitigation_rationale,
                residual_p1a, residual_p1b_p1c, residual_p1_sources,
                residual_p1, residual_hazardous_situation, severity, residual_poh
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            records,
        )

    # Also import Common Causes if present
    common_count = 0
    wb2 = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    if "Common Causes" in wb2.sheetnames:
        ws_cc = wb2["Common Causes"]
        cc_rows = list(ws_cc.iter_rows(values_only=True))
        if len(cc_rows) > 1:
            for row in cc_rows[1:]:
                module = s(row[0]) if len(row) > 0 else ""
                if not module:
                    continue
                db.execute(
                    "INSERT INTO fmea_common_causes (module, common_cause, common_cause_failure, mitigation, reference) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (
                        module,
                        s(row[1]) if len(row) > 1 else "",
                        s(row[2]) if len(row) > 2 else "",
                        s(row[3]) if len(row) > 3 else "",
                        s(row[4]) if len(row) > 4 else "",
                    ),
                )
                common_count += 1
    wb2.close()

    db.commit()

    return {
        "status": "success",
        "fmea_records": len(records),
        "common_causes": common_count,
        "sheet_used": target,
    }
